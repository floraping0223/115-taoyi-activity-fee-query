import csv
import json
from pathlib import Path

import openpyxl


SOURCE = Path(r"C:\Users\USER\Desktop\115點名名單.xlsx")
DATA_JS = Path("data.js")
OUTPUT_DIR = Path("outputs")
FIELD_REPORT = OUTPUT_DIR / "欄位對照與待補資料.md"
MISSING_CSV = OUTPUT_DIR / "待補資料.csv"


def text(value):
    return "" if value is None else str(value).strip()


def infer_group(value):
    label = text(value)
    if "小蟻" in label or any(name in label for name in ["小黑蟻", "小黃蟻", "小綠蟻", "小紅蟻"]):
        return "小蟻"
    if "炫蜂" in label or any(name in label for name in ["泥壺蜂", "虎頭蜂", "長腳蜂", "細腰蜂"]):
        return "炫蜂"
    if "奔鹿" in label or any(name in label for name in ["高地鹿", "森林鹿", "草原鹿", "湖泊鹿"]):
        return "奔鹿"
    if "翔鷹" in label or label == "鷹" or "鷹團" in label:
        return "翔鷹"
    return ""


def normalize_group(raw_group, role, raw_squad=""):
    value = text(raw_group)
    squad = text(raw_squad)
    source_group = infer_group(value)
    squad_group = infer_group(squad)
    if role == "成人" and "育成會" not in value:
        return source_group or squad_group or "育成會"
    if role == "成人":
        return "育成會"
    if source_group or squad_group:
        return source_group or squad_group
    return value or "未分團"


def normalize_role(raw_attr):
    value = text(raw_attr)
    if "家長" in value or "成人" in value:
        return "成人"
    if "小孩" in value:
        return "孩子"
    return value or "待補"


def normalize_squad(raw_squad):
    value = text(raw_squad)
    aliases = {
        "花叢-": "花叢",
        "天空-": "天空",
        "草原-": "草原",
        "大地-": "大地",
    }
    return aliases.get(value, value)


def main():
    workbook = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [text(cell) for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    mapping = {
        headers[0]: "家庭編號",
        headers[1]: "自然名",
        headers[2]: "屬性",
        headers[3]: "所屬分團",
        headers[4]: "所屬小隊",
    }

    people = []
    missing = []
    seen = set()
    for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        values = [text(value) for value in row[:5]]
        if not any(values):
            continue
        family_id, natural_name, raw_attr, source_group, source_squad = values
        if not source_squad:
            missing.append([row_number, family_id, natural_name, "所屬小隊", "已排除：不列入家庭確認、現場點名與後端統計"])
            continue
        squad = normalize_squad(source_squad)
        role = normalize_role(raw_attr)
        group = normalize_group(source_group, role, source_squad)
        person_id = f"P{row_number:04d}"
        person = {
            "id": person_id,
            "row": row_number,
            "familyId": family_id,
            "name": natural_name,
            "role": role,
            "group": group,
            "squad": squad,
            "sourceAttribute": raw_attr,
            "sourceGroup": source_group,
            "sourceSquad": source_squad,
            "eagleQualified": role == "成人" and group == "育成會" and "鷹" in source_group,
            "enabled": True,
        }
        key = (family_id, natural_name, role)
        if key in seen:
            missing.append([row_number, family_id, natural_name, "重複人員", "請確認是否同名或重複列"])
        seen.add(key)
        for field_name, field_value in [
            ("家庭編號", family_id),
            ("自然名", natural_name),
            ("屬性", raw_attr),
            ("所屬分團", source_group),
            ("所屬小隊", source_squad),
        ]:
            if not field_value or field_value == "/":
                missing.append([row_number, family_id, natural_name, field_name, "可運作，但建議補齊以利統計/分組"])
        if role == "待補":
            missing.append([row_number, family_id, natural_name, "屬性", "無法判斷成人或孩子"])
        people.append(person)

    OUTPUT_DIR.mkdir(exist_ok=True)
    js = "window.PEOPLE_DATA = " + json.dumps(people, ensure_ascii=False, indent=2) + ";\n"
    if DATA_JS.exists():
        old = DATA_JS.read_text(encoding="utf-8")
        marker = "window.ACTIVITY_FEE_DATA = "
        if marker in old:
            js += "\n" + old[old.index(marker):]
    DATA_JS.write_text(js, encoding="utf-8")

    with MISSING_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Excel列", "家庭編號", "自然名", "缺少/疑問欄位", "影響"])
        writer.writerows(missing)

    report_lines = [
        "# 欄位對照與待補資料",
        "",
        f"來源檔案：{SOURCE}",
        f"工作表：{sheet.title}",
        f"匯入人員數：{len(people)}",
        f"待補/疑問筆數：{len(missing)}",
        "",
        "## 欄位對照",
        "",
        "| 原始Excel欄位 | 系統欄位 | 處理方式 |",
        "| --- | --- | --- |",
    ]
    for source, target in mapping.items():
        report_lines.append(f"| {source.replace(chr(10), '<br>')} | {target} | 直接匯入/正規化 |")
    report_lines.extend([
        "",
        "## 育成鷹團資格判斷",
        "",
        "成人且「所屬分團」文字包含「鷹」即視為可於分流場次選擇育成鷹團活動。",
        "",
        "## 待補資料",
        "",
        "完整清單另存於 `outputs/待補資料.csv`。",
        "",
        "| Excel列 | 家庭編號 | 自然名 | 缺少/疑問欄位 | 影響 |",
        "| --- | --- | --- | --- | --- |",
    ])
    for item in missing[:60]:
        report_lines.append("| " + " | ".join(text(value) for value in item) + " |")
    if len(missing) > 60:
        report_lines.append(f"| ... | ... | ... | ... | 尚有 {len(missing) - 60} 筆，請見 CSV |")
    FIELD_REPORT.write_text("\n".join(report_lines) + "\n", encoding="utf-8")

    print(json.dumps({"people": len(people), "missing": len(missing), "headers": headers}, ensure_ascii=False))


if __name__ == "__main__":
    main()
